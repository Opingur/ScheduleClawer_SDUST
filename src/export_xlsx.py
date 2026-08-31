from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


WEEKDAYS = ("星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日")
PALETTE = ("F9B8C9", "A9CEF5", "BCE7D0", "E6C8F3", "F8D0A6", "AEE6E4", "F6D7A8", "C9D9F8")
BLUE = "2F6BDB"
DARK_BLUE = "1E4FA8"
LIGHT_BLUE = "EEF3FD"
BORDER = Border(
    left=Side(style="thin", color="D5DEEE"),
    right=Side(style="thin", color="D5DEEE"),
    top=Side(style="thin", color="D5DEEE"),
    bottom=Side(style="thin", color="D5DEEE"),
)


def _color_for(name: str) -> str:
    value = 0
    for char in name or "":
        value = (value * 31 + ord(char)) & 0xFFFFFFFF
    return PALETTE[value % len(PALETTE)]


def _card_text(item: dict) -> str:
    clock = (
        f"{item.get('startTime')}–{item.get('endTime')}"
        if item.get("startTime") and item.get("endTime")
        else f"第{item.get('startPeriod')}–{item.get('endPeriod')}节"
    )
    return "\n".join(
        str(value) for value in (
            item.get("courseName", ""),
            f"{clock}｜第{item.get('startPeriod')}–{item.get('endPeriod')}节",
            item.get("location", ""),
            item.get("teacher", ""),
        ) if value
    )


def _style_header(cells) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _style_grid(cell, fill: str | None = None) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _style_time_cell(cell) -> None:
    _style_grid(cell, LIGHT_BLUE)
    cell.font = Font(bold=True, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_title(sheet, title: str) -> None:
    sheet.merge_cells("A1:H1")
    cell = sheet["A1"]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.font = Font(bold=True, color="FFFFFF", size=15)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.sheet_view.showGridLines = False


def _set_widths(sheet, first: float, rest: float) -> None:
    sheet.column_dimensions["A"].width = first
    for column in "BCDEFGH":
        sheet.column_dimensions[column].width = rest


def _create_overview(book: Workbook, courses: list[dict], week_count: int) -> None:
    sheet = book.create_sheet("20周总览")
    _apply_title(sheet, "学期课表总览（按实际周次展开）")
    for column, value in enumerate(("周次", *WEEKDAYS), 1):
        sheet.cell(3, column, value)
    _style_header(sheet[3])
    for week in range(1, week_count + 1):
        row = week + 3
        cell = sheet.cell(row, 1, f"第{week}周")
        _style_time_cell(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        height = 70
        for weekday in range(1, 8):
            items = [item for item in courses if item["week"] == week and item["weekday"] == weekday]
            cell = sheet.cell(row, weekday + 1, "\n\n".join(_card_text(item) for item in items))
            _style_grid(cell, _color_for(items[0]["courseName"]) if items else None)
            height = max(height, min(180, 34 + len(items) * 62))
        sheet.row_dimensions[row].height = height
    _set_widths(sheet, 10, 27)
    sheet.freeze_panes = "B4"


def _date_label(term_start: date, week: int, weekday: int) -> str:
    current = term_start + timedelta(days=(week - 1) * 7 + (weekday - 1))
    return f"{current.month}月{current.day}日"


def _create_week_sheet(book: Workbook, courses: list[dict], week: int, term_start: date) -> None:
    sheet = book.create_sheet(f"第{week}周")
    _apply_title(sheet, f"第{week}周课表")
    for column, value in enumerate(("日期", *(_date_label(term_start, week, day) for day in range(1, 8))), 1):
        sheet.cell(3, column, value)
    _style_header(sheet[3])
    for column, value in enumerate(("时间 / 节次", *WEEKDAYS), 1):
        sheet.cell(4, column, value)
    _style_header(sheet[4])
    current = [item for item in courses if item["week"] == week]
    slots = {}
    for item in current:
        key = (item["startPeriod"], item["endPeriod"], item.get("startTime", ""), item.get("endTime", ""))
        slots.setdefault(key, item)
    if not slots:
        sheet.merge_cells("A5:H5")
        cell = sheet["A5"]
        cell.value = "本周没有识别到课程"
        cell.fill = PatternFill("solid", fgColor="F7F9FC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[5].height = 50
    for offset, slot in enumerate(sorted(slots.values(), key=lambda item: (item["startPeriod"], item["endPeriod"])), 5):
        clock = (
            f"{slot.get('startTime')}–{slot.get('endTime')}"
            if slot.get("startTime") and slot.get("endTime") else "未识别钟点"
        )
        time_cell = sheet.cell(offset, 1, f"{clock}\n第{slot['startPeriod']}–{slot['endPeriod']}节")
        _style_time_cell(time_cell)
        height = 80
        for weekday in range(1, 8):
            items = [
                item for item in current
                if item["weekday"] == weekday
                and item["startPeriod"] == slot["startPeriod"]
                and item["endPeriod"] == slot["endPeriod"]
            ]
            cell = sheet.cell(offset, weekday + 1, "\n\n".join(_card_text(item) for item in items))
            _style_grid(cell, _color_for(items[0]["courseName"]) if items else None)
            height = max(height, min(180, 34 + len(items) * 62))
        sheet.row_dimensions[offset].height = height
    _set_widths(sheet, 16, 24)
    sheet.freeze_panes = "B5"


def _create_details(book: Workbook, courses: list[dict]) -> None:
    sheet = book.create_sheet("课程明细")
    headers = ("周次", "星期", "课程名称", "具体时间", "节次", "教师", "地点", "教务系统时间原文")
    sheet.append(headers)
    _style_header(sheet[1])
    for item in courses:
        clock = (
            f"{item.get('startTime')}–{item.get('endTime')}"
            if item.get("startTime") and item.get("endTime") else "未识别钟点"
        )
        sheet.append((
            f"第{item['week']}周", WEEKDAYS[item["weekday"] - 1], item["courseName"], clock,
            f"第{item['startPeriod']}–{item['endPeriod']}节", item.get("teacher", ""),
            item.get("location", ""), item.get("sourceTime", ""),
        ))
    for row in sheet.iter_rows(min_row=2, max_row=max(2, sheet.max_row), max_col=8):
        for cell in row:
            _style_grid(cell)
    widths = (10, 10, 26, 16, 13, 16, 28, 30)
    for column, width in zip("ABCDEFGH", widths):
        sheet.column_dimensions[column].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 28
    if sheet.max_row >= 2:
        table = Table(displayName="CourseDetails", ref=f"A1:H{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"


def export_schedule(payload: dict, output: Path) -> None:
    courses = list(payload.get("courses") or [])
    week_count = int(payload.get("weekCount") or 20)
    try:
        term_start = date.fromisoformat(str(payload.get("termStartMonday") or "2026-08-31"))
    except ValueError as error:
        raise RuntimeError("第 1 周周一日期无效。") from error
    book = Workbook()
    book.remove(book.active)
    _create_overview(book, courses, week_count)
    for week in range(1, week_count + 1):
        _create_week_sheet(book, courses, week, term_start)
    _create_details(book, courses)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)